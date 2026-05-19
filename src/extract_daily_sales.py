import re
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

IN_PATH = Path("Data_raw_daily_sales_reports/daily_sales_2025.xlsx")
OUT_PATH = Path("data/processed/daily_sales_2025.csv")

DATE_RE = re.compile(r"^\s*(\d{1,2})/(\d{1,2})\s*$", re.IGNORECASE)

def as_number(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(",", "").replace("$", "")
    if s in {"", "-", "--"}:
        return None
    try:
        return float(s)
    except:
        return None

def as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str) and DATE_RE.match(value.strip()):
        month, day = map(int, value.strip().split("/"))
        return datetime(2025, month, day).date()
    return None

def extract_bread_projections(ws):
    projections = {}

    for r in range(1, ws.max_row + 1):
        am_header = ws.cell(r, 16).value
        pm_header = ws.cell(r, 18).value
        if not (
            isinstance(am_header, str)
            and isinstance(pm_header, str)
            and am_header.strip().upper() == "AM"
            and pm_header.strip().upper() == "PM"
        ):
            continue

        for rr in range(r + 1, min(r + 8, ws.max_row + 1)):
            d = as_date(ws.cell(rr, 15).value)
            if d is None:
                continue

            am_projection = as_number(ws.cell(rr, 16).value)
            in_process_by_11 = as_number(ws.cell(rr, 17).value)
            pm_projection = as_number(ws.cell(rr, 18).value)

            projections[d.isoformat()] = {
                "bread_am_projection": am_projection,
                "bread_in_process_by_11": in_process_by_11,
                "bread_pm_projection": pm_projection,
                "bread_total_projection": (
                    am_projection + pm_projection
                    if am_projection is not None and pm_projection is not None
                    else None
                ),
            }

    return projections

def main():
    if not IN_PATH.exists():
        raise FileNotFoundError(f"Input file not found: {IN_PATH.resolve()}")

    wb = load_workbook(IN_PATH, data_only=True)
    rows_out = []

    for ws in [wb["Projection Sheet"]]:
        # Read the whole sheet.
        max_r = ws.max_row
        max_c = ws.max_column
        bread_projections = extract_bread_projections(ws)

        # Find rows that contain dates like 12/17, 12/18, etc.
        for r in range(1, max_r + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, max_c + 1)]
            date_cols = []
            for c, v in enumerate(row_vals, start=1):
                if isinstance(v, str) and DATE_RE.match(v.strip()):
                    date_cols.append((c, v.strip()))
                elif isinstance(v, datetime):
                    # Sometimes Excel stores a real date value.
                    date_cols.append((c, v))

            # A typical week has several dates in the same row.
            if len(date_cols) < 5:
                continue

            # Find the nearby "Total" row below the date row.
            total_row = None
            for rr in range(r, min(r + 25, max_r) + 1):
                cell_c = ws.cell(rr, 3).value
                if isinstance(cell_c, str) and cell_c.strip().lower() == "total":
                    total_row = rr
                    break

            if total_row is None:
                continue

            # Extract sales for each date column.
            for c, dv in date_cols:
                sales_val = as_number(ws.cell(total_row, c).value)
                if sales_val is None:
                    continue

                # Build a real date; this source file is for 2025.
                d = as_date(dv)
                if d is None:
                    continue

                row = {"date": d.isoformat(), "total_sales": sales_val, "sheet": ws.title}
                row.update(bread_projections.get(d.isoformat(), {}))
                rows_out.append(row)

    if not rows_out:
        raise RuntimeError("No rows extracted. Check that the 'Total' row exists and dates are formatted like 12/17, 12/18, etc.")

    df = pd.DataFrame(rows_out).drop_duplicates(subset=["date"]).sort_values("date").reset_index(drop=True)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(OUT_PATH, index=False)

    print(f"Saved: {OUT_PATH}")
    print(df.head(10))
    print("rows=", len(df), "min_date=", df.date.min(), "max_date=", df.date.max())

if __name__ == "__main__":
    main()
