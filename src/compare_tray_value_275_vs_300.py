import math
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

IN_PATH = Path("Data_raw_daily_sales_reports/daily_sales_2025.xlsx")
WEEKLY_PATH = Path("data/processed/bread_weekly_2025.csv")
OUT_PATH = Path("data/processed/tray_value_275_vs_300_2025.csv")

FROZEN_TRAY_BREADS = 12
TRAY_VALUES = [275, 300]
PERCENT_AM = 1.0
PERCENT_PM = 0.5


def as_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("$", "")
    if text in {"", "-", "--"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def as_date(value):
    if isinstance(value, datetime):
        return value.date()
    return None


def projected_trays(projected_sales, tray_value, percent):
    if projected_sales is None:
        return None
    return math.ceil((projected_sales / tray_value) * percent)


def extract_projected_sales(ws):
    rows = []

    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 3).value
        if not isinstance(label, str) or label.strip().lower() != "am projected":
            continue

        pm_row = r + 1
        pm_label = ws.cell(pm_row, 3).value
        if not isinstance(pm_label, str) or pm_label.strip().lower() != "pm projected":
            continue

        date_row = r - 7
        for c in range(4, 11):
            date_value = as_date(ws.cell(date_row, c).value)
            if date_value is None:
                continue

            rows.append(
                {
                    "date": date_value.isoformat(),
                    "am_projected_sales": as_number(ws.cell(r, c).value),
                    "pm_projected_sales": as_number(ws.cell(pm_row, c).value),
                }
            )

    return pd.DataFrame(rows).drop_duplicates(subset=["date"]).sort_values("date")


def main():
    wb = load_workbook(IN_PATH, data_only=True)
    ws = wb["Projection Sheet"]

    daily = extract_projected_sales(ws)
    for tray_value in TRAY_VALUES:
        am_col = f"am_trays_{tray_value}"
        pm_col = f"pm_trays_{tray_value}"
        total_col = f"projected_breads_{tray_value}"

        daily[am_col] = daily["am_projected_sales"].apply(
            lambda value: projected_trays(value, tray_value, PERCENT_AM)
        )
        daily[pm_col] = daily["pm_projected_sales"].apply(
            lambda value: projected_trays(value, tray_value, PERCENT_PM)
        )
        daily[total_col] = (daily[am_col] + daily[pm_col]) * FROZEN_TRAY_BREADS

    daily["date"] = pd.to_datetime(daily["date"])
    start_date = pd.Timestamp("2025-01-01")
    daily = daily[
        (daily["date"] >= start_date)
        & (daily["date"] < start_date + pd.Timedelta(days=7 * 40))
    ].copy()
    daily["week"] = ((daily["date"] - start_date).dt.days // 7) + 1

    weekly_projection = (
        daily.groupby("week", as_index=False)
        .agg(
            projected_breads_275=("projected_breads_275", "sum"),
            projected_breads_300=("projected_breads_300", "sum"),
        )
    )
    weekly_projection["difference_300_minus_275"] = (
        weekly_projection["projected_breads_300"]
        - weekly_projection["projected_breads_275"]
    )

    weekly = pd.read_csv(WEEKLY_PATH)
    comparison = weekly_projection.merge(weekly, on="week", how="inner")

    for tray_value in TRAY_VALUES:
        projected_col = f"projected_breads_{tray_value}"
        comparison[f"projected_vs_actual_{tray_value}"] = (
            comparison[projected_col] - comparison["actual_usage"]
        )
        comparison[f"projected_vs_theoretical_{tray_value}"] = (
            comparison[projected_col] - comparison["theoretical_usage"]
        )

    comparison["unused_breads_from_usage_variance"] = -comparison["usage_variance"]
    comparison["unused_breads_$"] = -comparison["usage_variance_$"]

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    comparison.to_csv(OUT_PATH, index=False)

    print(f"Saved: {OUT_PATH}")
    print(f"Weeks compared: {len(comparison)}")
    print(
        "Average projected breads with 275: "
        f"{comparison['projected_breads_275'].mean():.2f}"
    )
    print(
        "Average projected breads with 300: "
        f"{comparison['projected_breads_300'].mean():.2f}"
    )
    print(
        "Average weekly difference, 300 minus 275: "
        f"{comparison['difference_300_minus_275'].mean():.2f}"
    )
    print(
        "Average projected vs theoretical, 275: "
        f"{comparison['projected_vs_theoretical_275'].mean():.2f}"
    )
    print(
        "Average projected vs theoretical, 300: "
        f"{comparison['projected_vs_theoretical_300'].mean():.2f}"
    )
    print(
        "Average unused breads from usage variance: "
        f"{comparison['unused_breads_from_usage_variance'].mean():.2f}"
    )
    print(f"Total unused bread cost: ${comparison['unused_breads_$'].sum():.2f}")
    print("\nFirst 10 weeks:")
    print(
        comparison[
            [
                "week",
                "projected_breads_275",
                "projected_breads_300",
                "difference_300_minus_275",
                "actual_usage",
                "theoretical_usage",
                "usage_variance",
                "usage_variance_$",
            ]
        ]
        .head(10)
        .to_string(index=False)
    )


if __name__ == "__main__":
    main()
